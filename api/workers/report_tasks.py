"""
Report Generation Tasks
=======================

Celery tasks for generating reports.
"""

import os
from datetime import datetime, timedelta, timezone
from pathlib import Path

from .celery_app import celery_app


@celery_app.task(bind=True)
def generate_report_task(self, report_id: int):
    """
    Generate a report document.
    
    Supports PDF and Excel formats.
    """
    import asyncio
    from sqlalchemy import select
    from api.core.database import get_db_context
    from api.models.analytics import Report
    
    async def run():
        async with get_db_context() as db:
            # Load report config
            result = await db.execute(
                select(Report).where(Report.id == report_id)
            )
            report = result.scalar_one_or_none()
            
            if not report:
                raise ValueError(f"Report {report_id} not found")
            
            try:
                # Generate report based on type
                output_dir = Path("reports")
                output_dir.mkdir(parents=True, exist_ok=True)
                
                if report.file_format == "pdf":
                    file_path = generate_pdf_report(report, output_dir)
                elif report.file_format == "xlsx":
                    file_path = generate_excel_report(report, output_dir)
                else:
                    file_path = generate_csv_report(report, output_dir)
                
                # Update report record
                report.status = "completed"
                report.file_path = str(file_path)
                report.file_size = os.path.getsize(file_path)
                
                await db.commit()
                
                return {"status": "completed", "file_path": str(file_path)}
                
            except Exception as e:
                report.status = "failed"
                report.error_message = str(e)
                await db.commit()
                raise
    
    return asyncio.get_event_loop().run_until_complete(run())


def generate_pdf_report(report, output_dir: Path) -> Path:
    """Generate PDF report using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table
    from reportlab.lib.styles import getSampleStyleSheet
    
    file_path = output_dir / f"report_{report.id}_{datetime.now().strftime('%Y%m%d')}.pdf"
    
    doc = SimpleDocTemplate(str(file_path), pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    story.append(Paragraph(report.title, styles["Title"]))
    story.append(Spacer(1, 20))
    
    # Date range
    story.append(Paragraph(
        f"Period: {report.start_date.strftime('%Y-%m-%d')} to {report.end_date.strftime('%Y-%m-%d')}",
        styles["Normal"]
    ))
    story.append(Spacer(1, 20))
    
    # Description
    if report.description:
        story.append(Paragraph(report.description, styles["Normal"]))
        story.append(Spacer(1, 20))
    
    # Summary section
    story.append(Paragraph("Executive Summary", styles["Heading1"]))
    story.append(Paragraph(
        "This report provides an analysis of passenger detection data for the specified period.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 20))
    
    # Sample data table
    data = [
        ["Metric", "Value"],
        ["Total Passengers", "N/A"],
        ["Peak Occupancy", "N/A"],
        ["Average Dwell Time", "N/A"],
    ]
    table = Table(data)
    story.append(table)
    
    doc.build(story)
    
    return file_path


def generate_excel_report(report, output_dir: Path) -> Path:
    """Generate Excel report using openpyxl."""
    try:
        from openpyxl import Workbook
    except ImportError:
        # Fallback to CSV
        return generate_csv_report(report, output_dir)
    
    file_path = output_dir / f"report_{report.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    
    # Headers
    ws["A1"] = "Bus Vision Report"
    ws["A2"] = f"Period: {report.start_date.strftime('%Y-%m-%d')} to {report.end_date.strftime('%Y-%m-%d')}"
    
    # Data headers
    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    
    # Sample data
    ws["A5"] = "Total Passengers"
    ws["B5"] = "N/A"
    ws["A6"] = "Peak Occupancy"
    ws["B6"] = "N/A"
    
    wb.save(file_path)
    
    return file_path


def generate_csv_report(report, output_dir: Path) -> Path:
    """Generate CSV report."""
    import csv
    
    file_path = output_dir / f"report_{report.id}_{datetime.now().strftime('%Y%m%d')}.csv"
    
    with open(file_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Report", report.title])
        writer.writerow(["Start Date", report.start_date.isoformat()])
        writer.writerow(["End Date", report.end_date.isoformat()])
        writer.writerow(["Total Passengers", "N/A"])
        writer.writerow(["Peak Occupancy", "N/A"])
    
    return file_path


@celery_app.task
def generate_daily_report_task():
    """
    Generate automated daily report.
    
    Runs at midnight via Celery Beat.
    """
    import asyncio
    from sqlalchemy import select, func
    from api.core.database import get_db_context
    from api.models.job import DetectionJob
    from api.models.analytics import Report
    
    async def run():
        async with get_db_context() as db:
            # Yesterday's date range
            today = datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0)
            yesterday = today - timedelta(days=1)
            
            # Check if report already exists
            result = await db.execute(
                select(Report).where(
                    Report.report_type == "daily",
                    Report.start_date == yesterday,
                    Report.end_date == today
                )
            )
            if result.scalar_one_or_none():
                return {"status": "already_exists"}
            
            # Count jobs from yesterday
            jobs_result = await db.execute(
                select(func.count(DetectionJob.id)).where(
                    DetectionJob.created_at >= yesterday,
                    DetectionJob.created_at < today
                )
            )
            jobs_count = jobs_result.scalar() or 0
            
            # Create report record
            report = Report(
                user_id=1,  # System user
                report_type="daily",
                title=f"Daily Report - {yesterday.strftime('%Y-%m-%d')}",
                start_date=yesterday,
                end_date=today,
                config={"jobs_count": jobs_count},
                is_scheduled=True,
            )
            
            db.add(report)
            await db.commit()
            await db.refresh(report)
            
            # Generate the report
            generate_report_task.delay(report.id)
            
            return {"status": "scheduled", "report_id": report.id}
    
    return asyncio.get_event_loop().run_until_complete(run())
